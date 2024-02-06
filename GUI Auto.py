import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import sv_ttk
import threading
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

# Global variables
selected_file = None
selected_sheet = None

# UI element for sheet selection
sheet_dropdown = None

def update_sheet_names():
    global selected_file, sheet_dropdown
    if selected_file:
        try:
            wb = load_workbook(selected_file, read_only=True)
            sheet_names = wb.sheetnames
            sheet_dropdown['values'] = sheet_names
            if sheet_names:
                sheet_dropdown.current(0)  # Set to first sheet as default
        except Exception as e:
            messagebox.showerror("Error", "Failed to read sheet names: " + str(e))

def select_file():
    global selected_file
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if file_path:
        selected_file = file_path
        update_sheet_names()

# Indices of columns to be filtered
## Create a dictionary to store values to be filtered
COLUMN_INDICES = {
    "N": 14, # 'N' is in column 'O' = 14
    "SPiCEinter_2thr": 129, # 'SPiCEinter_2thr' is in column 'DZ' = 129
    "Informe": 15, # 'Informe' is in column 'P' = 15
    "OMIM": 26, # 'OMIM' is in column 'AA' = 26
    "gnomad_HetCount_all": 92, # 'gnomad_HetCount_all" is in column 'CO' = 92
    "gnomad_AF_afr": 84, # 'gnomad_AF_afr' is in column 'CG' = 84
    "gnomad_AF_amr": 85, # 'gnomad_AF_amr' is in column 'CH' = 85
    "gnomad_AF_eas": 87, # 'gnomad_AF_eas' is in column 'CJ' = 87
    "gnomad_AF_sas": 88, # 'gnomad_AF_sas' is in column 'CK' = 88
    "gnomad_AF_nfe": 89, # 'gnomad_AF_nfe' is in column 'CL' = 89
    "gnomad_AF_fin": 90, # 'gnomad_AF_fin' is in column 'CM' = 90
    "ClinVar": 28, # 'ClinVar" is in column 'AC' = 28
    "anonimous_ANNOTATION": 45, # 'anonimous_ANNOTATION' is in column 'AT' = 45
    "anonimous_Func_refGene": 46, # 'annonimous_Func_refGene" is in column 'AU' = 46
    "distNearestSS": 54, # 'distNearestSS' is in column 'BC' = 54
    "Emedgene_Tag": 10, # 'Emedgene_Tag' is in column 'K' = 10
    "Emedgene_Evidence_Text": 13, # 'Emedgene_Evidence_Text' is in column 'N' = 13
    "incidental_findings": 66, # 'incidental_findings' is in column 'BO' = 66
}

# Style for the filters
STYLES = {
    "GREEN_STYLE": {
        "fill": PatternFill(start_color='c6efce', end_color='c6efce', fill_type='solid'),
        "font": Font(color='006100')
    },
    "GREY_ORANGE_STYLE": {
        "fill": PatternFill(start_color="f2f2f2", end_color="f2f2f2", fill_type='solid'),
        "font": Font(bold=True, color='fa7d00')
    }
}

# Function to set value and style for a cell using style sets
def set_cell_style(cell, value, style_set):
    cell.value = value
    cell.fill = style_set["fill"]
    cell.font = style_set["font"]

# Function to run the script
def run_script():
    if not selected_file or not selected_sheet:
        messagebox.showwarning("Warning", "Please select a file and a sheet.")
        return
    
    try:
        wb = load_workbook(selected_file)
        ws = wb[selected_sheet]
        
        ## Loop through and apply filters (1st round "Green")
        for row in ws.iter_rows(min_row=2):
            ### 20
            if (row[COLUMN_INDICES["N"]].value == 20) and \
            (row[COLUMN_INDICES["SPiCEinter_2thr"]].value in ['low', 'Outside SPiCE Interpretation']) and \
            (not row[COLUMN_INDICES["Informe"]].value):
                set_cell_style(row[COLUMN_INDICES["Informe"]], 'No x20', STYLES["GREEN_STYLE"])
            ### AD
            if ("recessive" not in str(row[COLUMN_INDICES["OMIM"]].value)) and \
            (row[COLUMN_INDICES["gnomad_HetCount_all"]].value > 10) and \
            (not row[COLUMN_INDICES["Informe"]].value):
                set_cell_style(row[COLUMN_INDICES["Informe"]], 'No x frec en AD', STYLES["GREEN_STYLE"])
            ### AR
            if ("recessive" in str(row[COLUMN_INDICES["OMIM"]].value)) and \
            (any(row[COLUMN_INDICES[col]].value > 0.002 for col in ["gnomad_AF_afr", "gnomad_AF_amr", "gnomad_AF_eas", "gnomad_AF_sas", "gnomad_AF_nfe", "gnomad_AF_fin"] if row[COLUMN_INDICES[col]].value is not None)) and \
            (not row[COLUMN_INDICES["Informe"]].value):
                set_cell_style(row[COLUMN_INDICES["Informe"]], 'No x frec en AR', STYLES["GREEN_STYLE"])


        ## Loop through and apply filters (2nd round "Grey")
        for row in ws.iter_rows(min_row=2):
            ### "Patho" 
            if ("Pathogenic" in str(row[COLUMN_INDICES["ClinVar"]].value) or "Likely_pathogenic" in str(row[COLUMN_INDICES["ClinVar"]].value)) and \
            (str(row[COLUMN_INDICES["Informe"]].value) not in ['No x20', 'No x frec en AD', 'No x frec en AR']):
                set_cell_style(row[COLUMN_INDICES["Informe"]], 'Patho', STYLES["GREY_ORANGE_STYLE"])

            ### "LOF" 
            if (any(term in str(row[COLUMN_INDICES["anonimous_ANNOTATION"]].value) for term in ["frameshift deletion", "frameshift insertion", "start_lost", "stopgain"])) and \
                (str(row[COLUMN_INDICES["Informe"]].value) not in ['No x20', 'No x frec en AD', 'No x frec en AR']) and \
                (str(row[COLUMN_INDICES["anonimous_ANNOTATION"]].value) not in ['nonframeshift deletion', 'nonframeshift insertion']):
                new_value = f'{row[COLUMN_INDICES["Informe"]].value}, LOF' if row[COLUMN_INDICES["Informe"]].value == 'Patho' else 'LOF'
                set_cell_style(row[COLUMN_INDICES["Informe"]], new_value, STYLES["GREY_ORANGE_STYLE"])

            ### "Splicing"
            if ("splice_acceptor_variant" in str(row[COLUMN_INDICES["anonimous_ANNOTATION"]].value) or "splice_donor_variant" in str(row[COLUMN_INDICES["anonimous_ANNOTATION"]].value)) and \
                ("splicing" in str(row[COLUMN_INDICES["anonimous_Func_refGene"]].value)) and \
                (row[COLUMN_INDICES["distNearestSS"]].value in [-2, -1, 1, 2]) and \
                (str(row[COLUMN_INDICES["Informe"]].value) not in ['No x20', 'No x frec en AD', 'No x frec en AR']):
                new_value = f'{row[COLUMN_INDICES["Informe"]].value}, Splicing' if any(term in str(row[COLUMN_INDICES["Informe"]].value) for term in ['Patho', 'LOF']) else 'Splicing'
                set_cell_style(row[COLUMN_INDICES["Informe"]], new_value, STYLES["GREY_ORANGE_STYLE"])

            ### "Most Likely"
            if ("most_likely" in str(row[COLUMN_INDICES["Emedgene_Tag"]].value)) and \
                (str(row[COLUMN_INDICES["Informe"]].value) not in ['No x20', 'No x frec en AD', 'No x frec en AR']):
                new_value = f'{row[COLUMN_INDICES["Informe"]].value}, Most Likely' if any(term in str(row[COLUMN_INDICES["Informe"]].value) for term in ['Patho', 'LOF', 'Splicing']) else 'Most Likely'
                set_cell_style(row[COLUMN_INDICES["Informe"]], new_value, STYLES["GREY_ORANGE_STYLE"])

            ### "Candidate"
            if ("candidate" in str(row[COLUMN_INDICES["Emedgene_Tag"]].value)) and \
                ("Match" in str(row[COLUMN_INDICES["Emedgene_Evidence_Text"]].value)) and \
                (str(row[COLUMN_INDICES["Informe"]].value) not in ['No x20', 'No x frec en AD', 'No x frec en AR']):
                new_value = f'{row[COLUMN_INDICES["Informe"]].value}, Candidate' if any(term in str(row[COLUMN_INDICES["Informe"]].value) for term in ['Patho', 'LOF', 'Splicing', 'Most Likely']) else 'Candidate'
                set_cell_style(row[COLUMN_INDICES["Informe"]], new_value, STYLES["GREY_ORANGE_STYLE"])

            ### "ACMG"
            if ("YES" in str(row[COLUMN_INDICES["incidental_findings"]].value)) and \
                (str(row[COLUMN_INDICES["Informe"]].value) not in ['No x20', 'No x frec en AD', 'No x frec en AR']):
                new_value = f'{row[COLUMN_INDICES["Informe"]].value}, ACMG' if any(term in str(row[COLUMN_INDICES["Informe"]].value) for term in ['Patho', 'LOF', 'Splicing', 'Most Likely', 'Candidate']) else 'ACMG'
                set_cell_style(row[COLUMN_INDICES["Informe"]], new_value, STYLES["GREY_ORANGE_STYLE"])

        # Save file
        wb.save('Processed_' + selected_file)

        # Message: Confirm successful implementation 
        messagebox.showinfo("Success", "Processing completed successfully.")

    except Exception as e:
        messagebox.showerror("Error", "An error occurred: " + str(e))

def start_script_thread():
    threading.Thread(target=run_script).start()

# Create the main window:
root = tk.Tk()
root.title("Excel Filter Bot")

# Add a file selection button
file_button = tk.Button(root, text="Select File", command=select_file)
file_button.pack()

# Sheet selection dropdown
sheet_dropdown = ttk.Combobox(root)
sheet_dropdown.pack()

# Run button
run_button = tk.Button(root, text="Run", command=lambda: threading.Thread(target=run_script).start())
run_button.pack()

sv_ttk.set_theme("light")

root.mainloop()
