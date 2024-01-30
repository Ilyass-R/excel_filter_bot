# Version with dictionaries and styling functions
# IMPORTANT: Before running this code make sure to have the Excel file in question closed.
# Test for unchanged filters

# Load required libraries
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

# Create a dictionary to store values to be filtered
COLUMN_INDICES = {
    "N": 14, # 'N' is in column 'O' = 14
    "SPiCEinter_2thr": 129, # 'SPiCEinter_2thr' is in column 'DZ' = 129
    "Informe": 15, # 'Informe' is in column 'P' = 15
    "OMIM": 26, # 'OMIM' is in column 'AA' = 26
    "gnomad_HetCount_all": 92, # 'gnomad_HetCount_all" is in column 'CO' = 92
    "gnomad_AF_afr": 84, # 'gnomad_AF_afr' is in column 'CG' = 84
    "gnomad_AF_amr": 85, # 'gnomad_AF_amr' is in column 'CH' = 85
    "gnomad_AF_eas": 87, # 'gnomad_AF_eas' is in column 'CJ' = 87
    "gnomad_AF_sas": 88, # 'gnomad_AF_sas' is in column 'CK' = 88
    "gnomad_AF_nfe": 89, # 'gnomad_AF_nfe' is in column 'CL' = 89
    "gnomad_AF_fin": 90, # 'gnomad_AF_fin' is in column 'CM' = 90
    "ClinVar": 28, # 'ClinVar" is in column 'AC' = 28
    "anonimous_ANNOTATION": 45, # 'anonimous_ANNOTATION' is in column 'AT' = 45
    "anonimous_Func_refGene": 46, # 'annonimous_Func_refGene" is in column 'AU' = 46
    "distNearestSS": 54, # 'distNearestSS' is in column 'BC' = 54
    "Emedgene_Tag": 10, # 'Emedgene_Tag' is in column 'K' = 10
    "Emedgene_Evidence_Text": 13, # 'Emedgene_Evidence_Text' is in column 'N' = 13
    "incidental_findings": 66, # 'incidental_findings' is in column 'BO' = 66
}

# Style for the filters
## First Round
GREEN_FILL = PatternFill(start_color='c6efce', end_color='c6efce', fill_type='solid')
GREEN_FONT = Font(color='006100')

## Second Round
GREY_FILL = PatternFill(start_color="f2f2f2", end_color="f2f2f2", fill_type='solid')
BOLD_ORANGE_FONT = Font(bold=True, color='fa7d00')

# Function to set value and style for a cell
def set_cell(cell, value, fill, font):
    cell.value = value
    cell.fill = fill
    cell.font = font

# Load Excel file and specific sheet
file_path = '23NR10186_latest_emedgene_destacado.xlsx'
wb = load_workbook(file_path)

# Check if the specific sheet exists
if 'OMIM+NIM CLIN+ClinVar' in wb.sheetnames:
    ws = wb['OMIM+NIM CLIN+ClinVar']

    ## Loop through and apply filters (1st round "Green")
    for row in ws.iter_rows(min_row=2):
        ### 'N' is in column 'O' = 14, 'SPiCEinter_2thr' is in column 'DZ' = 129 and 'Informe' is in column 'P' = 15, print 'No x20'
        ### Highlight cell in color "green"
        if (row[COLUMN_INDICES["N"]].value == 20) and \
            (row[COLUMN_INDICES["SPiCEinter_2thr"]].value in ['low', 'Outside SPiCE Interpretation']) and \
            (not row[COLUMN_INDICES["Informe"]].value):
            set_cell(row[COLUMN_INDICES["Informe"]], 'No x20', GREEN_FILL, GREEN_FONT)

        ### 'OMIM' is in column 'AA' = 26, 'gnomad_HetCount_all" is in column 'CO' = 92
        ### Print('No x frec en AD')
        if ("recessive" not in str(row[COLUMN_INDICES["OMIM"]].value)) and \
        (row[COLUMN_INDICES["gnomad_HetCount_all"]].value > 10) and \
        (not row[COLUMN_INDICES["Informe"]].value):
            set_cell(row[COLUMN_INDICES["Informe"]], 'No x frec en AD', GREEN_FILL, GREEN_FONT)

        ### 'OMIM' is in column 'AA' = 26, and gnomad_AF_afr is in column 'CG' = 84, gnomad_AF_amr 'CH' = 85, gnomad_AF_eas 'CJ' = 87, \
        ### gnomad_AF_sas 'CK' = 88, gnomad_AF_nfe 'CL' = 89, gnomad_AF_fin 'CM' = 90
        ### Print('No x frec en AR') 
        if ("recessive" in str(row[COLUMN_INDICES["OMIM"]].value)) and \
        (any(row[COLUMN_INDICES[col]].value > 0.002 for col in ["gnomad_AF_afr", "gnomad_AF_amr", "gnomad_AF_eas", "gnomad_AF_sas", "gnomad_AF_nfe", "gnomad_AF_fin"] if row[COLUMN_INDICES[col]].value is not None)) and \
        (not row[COLUMN_INDICES["Informe"]].value):
            set_cell(row[COLUMN_INDICES["Informe"]], 'No x frec en AR', GREEN_FILL, GREEN_FONT)


    ## Loop through and apply filters (2nd round "Grey")
    for row in ws.iter_rows(min_row=2):
        ### 'ClinVar" is in column 'AC' = 28, check if "Pathogenic" or "Likely_pathogenic" is in the cell, print "Patho"
        ### Highlight cell in color "grey" and set text style "bold" and color "orange"  
        if ("Pathogenic" in str(row[COLUMN_INDICES["ClinVar"]].value) or "Likely_pathogenic" in str(row[COLUMN_INDICES["ClinVar"]].value)) and \
        (str(row[COLUMN_INDICES["Informe"]].value) not in ['No x20', 'No x frec en AD', 'No x frec en AR']):
            set_cell(row[COLUMN_INDICES["Informe"]], 'Patho', GREY_FILL, BOLD_ORANGE_FONT)

        ### 'anonimous_ANNOTATION' is in column 'AT' = 45, check if "frameshift deletion", "frameshift insertion", or "start lost" or "stopgain" is in the cell
        ### If the previous cell has been filtered as "Patho": append "Frameshift" to the previous "Informe" cell, else print "Frameshift"
        ### Highlight cell in color "grey" and set text style "bold" and color "orange"  
        if (any(term in str(row[COLUMN_INDICES["anonimous_ANNOTATION"]].value) for term in ["frameshift deletion", "frameshift insertion", "start_lost", "stopgain"])) and \
            (str(row[COLUMN_INDICES["Informe"]].value) not in ['No x20', 'No x frec en AD', 'No x frec en AR']) and \
            (str(row[COLUMN_INDICES["anonimous_ANNOTATION"]].value) not in ['nonframeshift deletion', 'nonframeshift insertion']):
            new_value = f'{row[COLUMN_INDICES["Informe"]].value}, LOF' if row[COLUMN_INDICES["Informe"]].value == 'Patho' else 'LOF'
            set_cell(row[COLUMN_INDICES["Informe"]], new_value, GREY_FILL, BOLD_ORANGE_FONT)

        ### 'anonimous_ANNOTATION' is in column 'AT' = 45, check if it mentions "splice_acceptor_variant" or "splice_donor_variant"
        ### 'annonimous_Func_refGene" is in column 'AU' = 46, check if it mentions "splicing"
        ### 'distNearestSS' is in column 'BC' = 54, check if values: [(-2, -1, 1, 2)]
        ### If the previous cell has been filtered as "Patho" or "Frameshift" or "Patho, Frameshift": add "Splicing" to the "Informe" cell, else print "Splicing"
        ### Highlight cell in color "grey" and set text style "bold" and color "orange"
        if ("splice_acceptor_variant" in str(row[COLUMN_INDICES["anonimous_ANNOTATION"]].value) or "splice_donor_variant" in str(row[COLUMN_INDICES["anonimous_ANNOTATION"]].value)) and \
            ("splicing" in str(row[COLUMN_INDICES["anonimous_Func_refGene"]].value)) and \
            (row[COLUMN_INDICES["distNearestSS"]].value in [-2, -1, 1, 2]) and \
            (str(row[COLUMN_INDICES["Informe"]].value) not in ['No x20', 'No x frec en AD', 'No x frec en AR']):
            new_value = f'{row[COLUMN_INDICES["Informe"]].value}, Splicing' if any(term in str(row[COLUMN_INDICES["Informe"]].value) for term in ['Patho', 'LOF']) else 'Splicing'
            set_cell(row[COLUMN_INDICES["Informe"]], new_value, GREY_FILL, BOLD_ORANGE_FONT)

        ### 'Emedgene_Tag' is in column 'K' = 10, check if it mentions "most_likely"
        ### If the previous cell has been filtered as "Patho" or "Frameshift" or "Patho, Frameshift, Splicing": add "Most Likely" to the "Informe" cell, else print "Most Likely"
        ### Highlight cell in color "grey" and set text style "bold" and color "orange"
        if ("most_likely" in str(row[COLUMN_INDICES["Emedgene_Tag"]].value)) and \
            (str(row[COLUMN_INDICES["Informe"]].value) not in ['No x20', 'No x frec en AD', 'No x frec en AR']):
            new_value = f'{row[COLUMN_INDICES["Informe"]].value}, Most Likely' if any(term in str(row[COLUMN_INDICES["Informe"]].value) for term in ['Patho', 'LOF', 'Splicing']) else 'Most Likely'
            set_cell(row[COLUMN_INDICES["Informe"]], new_value, GREY_FILL, BOLD_ORANGE_FONT)

        ### 'Emedgene_Tag' is in column 'K' = 10, check if it mentions "candidate" 
        ###  AND 'Emedgene_Evidence_Text' is in column 'N' = 13, check if it mentions "Match"
        ### If the previous cell has been filtered as "Patho" or "Frameshift" or "Patho, Frameshift, Splicing, Most Likely": add "Candidate" to the "Informe" cell, else print "Candidate"
        ### Highlight cell in color "grey" and set text style "bold" and color "orange"
        if ("candidate" in str(row[COLUMN_INDICES["Emedgene_Tag"]].value)) and \
            ("Match" in str(row[COLUMN_INDICES["Emedgene_Evidence_Text"]].value)) and \
            (str(row[COLUMN_INDICES["Informe"]].value) not in ['No x20', 'No x frec en AD', 'No x frec en AR']):
            new_value = f'{row[COLUMN_INDICES["Informe"]].value}, Candidate' if any(term in str(row[COLUMN_INDICES["Informe"]].value) for term in ['Patho', 'LOF', 'Splicing', 'Most Likely']) else 'Candidate'
            set_cell(row[COLUMN_INDICES["Informe"]], new_value, GREY_FILL, BOLD_ORANGE_FONT)

        ### 'incidental_findings' is in column 'BO' = 66, check if it mentions "YES"
        ### If the previous cell has been filtered as "Patho" or "Frameshift" or "Patho, Frameshift, Splicing, Most Likely, Candidate" to the "Informe" cell, else print "Most Likely"
        ### Highlight cell in color "grey" and set text style "bold" and color "orange"
        if ("YES" in str(row[COLUMN_INDICES["incidental_findings"]].value)) and \
            (str(row[COLUMN_INDICES["Informe"]].value) not in ['No x20', 'No x frec en AD', 'No x frec en AR']):
            new_value = f'{row[COLUMN_INDICES["Informe"]].value}, ACMG' if any(term in str(row[COLUMN_INDICES["Informe"]].value) for term in ['Patho', 'LOF', 'Splicing', 'Most Likely', 'Candidate']) else 'ACMG'
            set_cell(row[COLUMN_INDICES["Informe"]], new_value, GREY_FILL, BOLD_ORANGE_FONT)

        ### 'OMIM' is in column 'AA' = 26, check if it mentions "ND"
        ### If the previous cell has been filtered as "Patho" or "Frameshift" or "Patho, Frameshift, Splicing, Most Likely, Candidate, ACMG" to the "Informe" cell, else print "No x ND"
        ### Highlight cell in color "grey" and set text style "bold" and color "orange"
        
        # if "ND" in str(row[26].value):
            # row[1].value = f'{row[1].value}, ND' if any(term in str(row[1].value) for term in ['Patho', 'LOF', 'Splicing', 'Most Likely', 'Candidate', 'ACMG']) else 'No por ND'
            # row[15].value = 'No x ND'
            # row[15].fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')

        #### Check entry 76 (Particular case)
        #### Make sure if the filter above applies sequentially or after all the above filters have been applied.
        #### Can ND be checked at the same time as the rest of the other filters or not?   

# Save file
new_file_path = 'F_Filtered_' + file_path.split('/')[-1]
wb.save(new_file_path)
