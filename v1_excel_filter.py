# Currently working test version
# IMPORTANT: Before running this code make sure to have the Excel file in question closed.
# Test for unchanged filters

# Load required libraries
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

# Load Excel file and specific sheet
file_path = '23NR10186_latest_emedgene_destacado.xlsx'
wb = load_workbook(file_path)

# Check if the specific sheet exists
if 'OMIM+NIM CLIN+ClinVar' in wb.sheetnames:
    ws = wb['OMIM+NIM CLIN+ClinVar']

    ## Loop through and apply filters (1st round "Green")
    for row in ws.iter_rows(min_row=2):
        ### 'N' is in column 'O' = 14, 'SPiCEinter_2thr' is in column 'DZ' = 129 and 'Informe' is in column 'P' = 15, print 'No x20'
        ### Highlight cell in color "green"
        if (row[14].value == 20) and \
            (row[129].value in ['low', 'Outside SPiCE Interpretation']) and \
            (not row[15].value):
            row[15].value = 'No x20'
            row[15].fill = PatternFill(start_color='c6efce', end_color='c6efce', fill_type='solid')
            row[15].font = Font(color='006100') 

        ### 'OMIM' is in column 'AA' = 26, 'gnomad_HetCount_all" is in column 'CO' = 92
        ### Print('No x frec en AD')
        if ("recessive" not in str(row[26].value)) and \
            ((row[92].value > 10)) and \
            (not row[15].value):
            row[15].value = 'No x frec en AD'
            row[15].fill = PatternFill(start_color='c6efce', end_color='c6efce', fill_type='solid')
            row[15].font = Font(color='006100')

        ### 'OMIM' is in column 'AA' = 26, and gnomad_AF_afr is in column 'CG' = 84, gnomad_AF_amr 'CH' = 85, gnomad_AF_eas 'CJ' = 87, \
        ### gnomad_AF_sas 'CK' = 88, gnomad_AF_nfe 'CL' = 89, gnomad_AF_fin 'CM' = 90
        ### Print('No x frec en AR') 
        if ("recessive" in str(row[26].value)) and \
            (any([row[i].value > 0.002 for i in [84, 85, 87, 88, 89, 90] if row[i].value is not None])) and \
            (not row[15].value):
            row[15].value = 'No x frec en AR'
            row[15].fill = PatternFill(start_color='c6efce', end_color='c6efce', fill_type='solid')
            row[15].font = Font(color='006100')

    ## Loop through and apply filters (2nd round "Grey")
    for row in ws.iter_rows(min_row=2):
        ### 'ClinVar" is in column 'AC' = 28, check if "Pathogenic" or "Likely_pathogenic" is in the cell, print "Patho"
        ### Highlight cell in color "grey" and set text style "bold" and color "orange"  
        if ("Pathogenic" in str(row[28].value) or "Likely_pathogenic" in str(row[28].value)) and \
            (str(row[15].value) not in ['No x20', 'No x frec en AD', 'No x frec en AR']):
            row[15].value = 'Patho'
            row[15].fill = PatternFill(start_color="f2f2f2", end_color="f2f2f2", fill_type='solid')
            row[15].font = Font(bold=True ,color='fa7d00')

        ### 'anonimous_ANNOTATION' is in column 'AT' = 45, check if "frameshift deletion", "frameshift insertion", or "start lost" or "stopgain" is in the cell
        ### If the previous cell has been filtered as "Patho": append "Frameshift" to the previous "Informe" cell, else print "Frameshift"
        ### Highlight cell in color "grey" and set text style "bold" and color "orange"  
        if (any(term in str(row[45].value) for term in ["frameshift deletion", "frameshift insertion", "start_lost", "stopgain"])) and \
            (str(row[15].value) not in ['No x20', 'No x frec en AD', 'No x frec en AR']) and \
            (str(row[45].value) not in ['nonframeshift deletion', 'nonframeshift insertion']):
            row[15].value = f'{row[15].value}, LOF' if row[15].value == 'Patho' else 'LOF'
            row[15].fill = PatternFill(start_color="f2f2f2", end_color="f2f2f2", fill_type='solid')
            row[15].font = Font(bold=True ,color='fa7d00')

        ### 'anonimous_ANNOTATION' is in column 'AT' = 45, check if it mentions "splice_acceptor_variant" or "splice_donor_variant"
        ### 'annonimous_Func_refGene" is in column 'AU' = 46, check if it mentions "splicing"
        ### 'distNearestSS' is in column 'BC' = 54, check if values: [(-2, -1, 1, 2)]
        ### If the previous cell has been filtered as "Patho" or "Frameshift" or "Patho, Frameshift": add "Splicing" to the "Informe" cell, else print "Splicing"
        ### Highlight cell in color "grey" and set text style "bold" and color "orange"
            
        if ("splice_acceptor_variant" in str(row[45].value) or "splice_donor_variant" in str(row[45].value)) and \
            ("splicing" in str(row[46].value)) and \
            (row[54].value in [-2, -1, 1, 2]) and \
            (str(row[15].value) not in ['No x20', 'No x frec en AD', 'No x frec en AR']):
                row[15].value = f'{row[15].value}, Splicing' if any(term in str(row[15].value) for term in ['Patho', 'LOF']) else 'Splicing'
                row[15].fill = PatternFill(start_color="f2f2f2", end_color="f2f2f2", fill_type='solid')
                row[15].font = Font(bold=True ,color='fa7d00')

        ### 'Emedgene_Tag' is in column 'K' = 10, check if it mentions "most_likely"
        ### If the previous cell has been filtered as "Patho" or "Frameshift" or "Patho, Frameshift, Splicing": add "Most Likely" to the "Informe" cell, else print "Most Likely"
        ### Highlight cell in color "grey" and set text style "bold" and color "orange"
        
        if ("most_likely" in str(row[10].value)) and \
            (str(row[15].value) not in ['No x20', 'No x frec en AD', 'No x frec en AR']):
            row[15].value = f'{row[15].value}, Most Likely' if any(term in str(row[15].value) for term in ['Patho', 'LOF', 'Splicing']) else 'Most Likely'
            row[15].fill = PatternFill(start_color="f2f2f2", end_color="f2f2f2", fill_type='solid')
            row[15].font = Font(bold=True ,color='fa7d00')

        ### 'Emedgene_Tag' is in column 'K' = 10, check if it mentions "candidate" 
        ###  AND 'Emedgene_Evidence_Text' is in column 'N' = 13, check if it mentions "Match"
        ### If the previous cell has been filtered as "Patho" or "Frameshift" or "Patho, Frameshift, Splicing, Most Likely": add "Candidate" to the "Informe" cell, else print "Candidate"
        ### Highlight cell in color "grey" and set text style "bold" and color "orange"
        
        if ("candidate" in str(row[10].value)) and \
            ("Match" in str(row[13].value)) and \
            (str(row[15].value) not in ['No x20', 'No x frec en AD', 'No x frec en AR']):
            row[15].value = f'{row[15].value}, Candidate' if any(term in str(row[15].value) for term in ['Patho', 'LOF', 'Splicing', 'Most Likely']) else 'Candidate'
            row[15].fill = PatternFill(start_color="f2f2f2", end_color="f2f2f2", fill_type='solid')
            row[15].font = Font(bold=True ,color='fa7d00')

        ### 'incidental_findings' is in column 'BO' = 66, check if it mentions "YES"
        ### If the previous cell has been filtered as "Patho" or "Frameshift" or "Patho, Frameshift, Splicing, Most Likely, Candidate" to the "Informe" cell, else print "Most Likely"
        ### Highlight cell in color "grey" and set text style "bold" and color "orange"
        
        if ("YES" in str(row[66].value)) and \
            (str(row[15].value) not in ['No x20', 'No x frec en AD', 'No x frec en AR']):
            row[15].value = f'{row[15].value}, ACMG' if any(term in str(row[15].value) for term in ['Patho', 'LOF', 'Splicing', 'Most Likely', 'Candidate']) else 'ACMG'
            row[15].fill = PatternFill(start_color="f2f2f2", end_color="f2f2f2", fill_type='solid')
            row[15].font = Font(bold=True ,color='fa7d00')

        ### 'OMIM' is in column 'AA' = 26, check if it mentions "ND"
        ### If the previous cell has been filtered as "Patho" or "Frameshift" or "Patho, Frameshift, Splicing, Most Likely, Candidate, ACMG" to the "Informe" cell, else print "No x ND"
        ### Highlight cell in color "grey" and set text style "bold" and color "orange"
        
        # if "ND" in str(row[26].value):
            # row[1].value = f'{row[1].value}, ND' if any(term in str(row[1].value) for term in ['Patho', 'LOF', 'Splicing', 'Most Likely', 'Candidate', 'ACMG']) else 'No por ND'
            # row[15].value = 'No x ND'
            # row[15].fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')

        #### Check entry 76 (Particular case)
        #### Make sure if the filter above applies sequentially or after all the above filters have been applied.
        #### Can ND be checked at the same time as the rest of the other filters or not?   

# Save file
new_file_path = 'Filtered_' + file_path.split('/')[-1]
wb.save(new_file_path)